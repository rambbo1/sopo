"""
엑셀 생성 모듈 — 기존 매출집계 파일과 동일한 형식으로 출력
환율 적용 기준: 소포수령증 발행일(write_date) 환율
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange
from typing import Optional
import re
from pathlib import Path


# ── 스타일 정의 ────────────────────────────────────────────────
HEADER_FILL   = PatternFill('solid', start_color='D9E1F2', end_color='D9E1F2')
SUBHEAD_FILL  = PatternFill('solid', start_color='E2EFDA', end_color='E2EFDA')
GRAY_FILL     = PatternFill('solid', start_color='F2F2F2', end_color='F2F2F2')

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)

FONT_DEFAULT = Font(name='맑은 고딕', size=9)
FONT_BOLD    = Font(name='맑은 고딕', size=9, bold=True)
FONT_TITLE   = Font(name='맑은 고딕', size=11, bold=True)

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
RIGHT  = Alignment(horizontal='right',  vertical='center')

# SMBS 환율은 일부 통화가 100단위 기준 (원화 = 외화 × 환율 / 제수)
RATE_DIVISOR = {
    'JPY': 100,   # 100엔 기준
    'VND': 100,   # 100동 기준
}

# 제출자(판매자) 정보 기본값 — PDF에서 못 읽었을 때만 사용
DEFAULT_SUBMITTER = {
    'name': '유엠(UM)', 'biz_no': '529-12-02268',
    'ceo': '맹진열', 'address': '서울특별시 광진구 광나루로 556, 1동 2층',
}

# 숫자 천단위 콤마 서식
NUM_FMT  = '#,##0'        # 정수(수량·원화)
NUM_FMT2 = '#,##0.00'     # 소수(외화·환율)

# L/C 번호 또는 수출신고번호에는 운송장번호를 넣지 않습니다.
# 영세율 증빙은 기타영세율건수 1로 신고합니다.
TRACKING_NO_PATTERN = re.compile(r"^[A-Z]{2}[A-Z0-9]{13}$", re.I)

def is_valid_tracking_no(value):
    text = re.sub(r"[^A-Za-z0-9]", "", str(value or "")).upper()
    return bool(TRACKING_NO_PATTERN.fullmatch(text))

def other_zero_rate_count_value(value=None):
    return 1

def _date_to_int(value):
    d = re.sub(r"\D", "", str(value or ""))[:8]
    return int(d) if len(d) == 8 else None



def _months_between(start, end):
    """거래기간(start~end)이 포함하는 (연,월) 리스트."""
    def _p(x):
        d = re.sub(r'\D', '', str(x))[:8]
        return (int(d[:4]), int(d[4:6])) if len(d) >= 6 else None
    a = _p(start); b = _p(end)
    a = a or b; b = b or a
    if not a:
        return []
    if b < a:
        a, b = b, a
    out = []
    y, m = a
    while (y, m) <= b:
        out.append((y, m))
        m += 1
        if m > 12:
            m = 1; y += 1
    return out


def period_labels(shopee_results, lazada_result, qoo10_result, ebay_results=None, fallback=''):
    """데이터 거래기간으로 (표시용, 파일명용) 라벨 생성.
    예: 10월 / 10~12월 / 1~12월 / 3,5,9월(파일명) · 3/5/9월(표시)."""
    pairs = []
    for sd in (shopee_results or []):
        pairs.append((sd.get('period_start', ''), sd.get('period_end', '')))
    if lazada_result:
        pairs.append((lazada_result.get('period_start', ''), lazada_result.get('period_end', '')))
    if qoo10_result:
        pairs.append((qoo10_result.get('period_start', ''), qoo10_result.get('period_end', '')))
    for er in (ebay_results or []):
        pairs.append((er.get('period_start', ''), er.get('period_end', '')))
    yms = set()
    for s_, e_ in pairs:
        yms.update(_months_between(s_, e_))
    if not yms:
        return fallback, fallback
    years = sorted(set(y for y, m in yms))

    def _fmt(y, list_sep):
        ms = sorted(m for yy, m in yms if yy == y)
        if len(ms) == 1:
            return f'{ms[0]}월'
        if ms == list(range(ms[0], ms[-1] + 1)):
            return f'{ms[0]}~{ms[-1]}월'
        return list_sep.join(str(m) for m in ms) + '월'

    if len(years) == 1:
        y = years[0]
        disp = f'{y}년 {_fmt(y, "/")}'
        fname = _fmt(y, ',')
    else:
        disp = ', '.join(f'{y}년 {_fmt(y, "/")}' for y in years)
        fname = '_'.join(f'{y}년{_fmt(y, ",")}' for y in years)
    return disp, fname


def _style(cell, font=None, fill=None, align=None, border=None, num_format=None):
    if font:      cell.font       = font
    if fill:      cell.fill       = fill
    if align:     cell.alignment  = align
    if border:    cell.border     = border
    if num_format: cell.number_format = num_format


# ── 소포수령증 표 열 그룹 (값 열 + 사이 빈 열을 병합해 깔끔하게 이어줌) ──
_RECEIPT_GROUPS_2 = [(1, 3), (4, 6), (7, 10), (11, 12), (13, 15), (16, 19)]
_RECEIPT_GROUPS_3 = [(1, 3), (4, 6), (7, 10), (11, 12), (13, 15), (16, 17), (18, 18), (19, 19)]


def _merge_row(ws, row, groups, border=None):
    """한 행에서 각 열 그룹을 병합하고(2칸 이상), 그룹 전체에 테두리를 적용.
    겹침 검사 없는 빠른 병합(대량 행 처리 속도 향상, 결과는 동일)."""
    for c1, c2 in groups:
        if c2 > c1:
            ws.merged_cells.ranges.add(
                CellRange(min_col=c1, min_row=row, max_col=c2, max_row=row))
        if border is not None:
            for col in range(c1, c2 + 1):
                ws.cell(row=row, column=col).border = border


def _get_rate(rates: dict, currency: str, date_str: str) -> float:
    """
    발행일(write_date) 기준 환율 반환.
    daily 데이터 없으면 average(수동입력값) 반환.
    date_str이 비어 있으면 average 반환.
    """
    from .exchange_rate import get_rate_for_date
    rate_data = rates.get(currency)
    if not rate_data:
        return 0.0
    # daily 데이터가 없으면 average 반환 (수동입력 모드)
    if not rate_data.get('daily'):
        return rate_data.get('average', 0.0)
    if not date_str:
        return rate_data.get('average', 0.0)
    rate = get_rate_for_date(rate_data, date_str)
    if rate == 0.0:
        rate = rate_data.get('average', 0.0)
    return rate


# ── 환율 시트 작성 ──────────────────────────────────────────────

def write_exchange_rate_sheet(ws, rate_data: dict):
    """환율(XXX) 시트를 SMBS 데이터로 채움"""
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12

    if rate_data is None:
        ws['A1'] = '환율 데이터 없음 (수동 입력 필요)'
        return

    # 월평균 전용 데이터만 있는 경우(이베이/린코스)
    if rate_data.get('monthly') and not rate_data.get('daily'):
        ws['A1'] = '월평균 매매기준율'
        _style(ws['A1'], font=FONT_BOLD)
        ws['A2'] = f"기간 : {rate_data.get('period', '')}"
        headers = ['년월', '통화명', '월평균환율']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=4, column=col, value=h)
            _style(c, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)
        for r, d in enumerate(rate_data.get('monthly', []), 5):
            vals = [d.get('year_month', ''), rate_data.get('currency_name', ''), round(float(d.get('rate', 0) or 0), 2)]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=r, column=col, value=v)
                _style(
                    c, font=FONT_DEFAULT, align=CENTER if col != 2 else LEFT, border=THIN_BORDER,
                    num_format=NUM_FMT2 if col == 3 else None,
                )
        return

    # 제목
    ws['A1'] = '기간별 매매기준율'
    _style(ws['A1'], font=FONT_BOLD)
    ws['A2'] = f"기간 : {rate_data['period']}"

    # 평균환율 통계
    ws['A4'] = '평균환율'
    _style(ws['A4'], font=FONT_BOLD)

    headers5 = ['평균환율', '최저치', '기록일', '최고치', '기록일', '등락폭', 'Cross Rate']
    for col, h in enumerate(headers5, 1):
        c = ws.cell(row=5, column=col, value=h)
        _style(c, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)

    vals6 = [
        rate_data.get('average', ''),
        rate_data.get('min', ''),
        rate_data.get('min_date', ''),
        rate_data.get('max', ''),
        rate_data.get('max_date', ''),
        rate_data.get('range', ''),
        rate_data.get('cross_rate', ''),
    ]
    for col, v in enumerate(vals6, 1):
        c = ws.cell(row=6, column=col, value=v)
        _style(
            c, font=FONT_DEFAULT, align=CENTER, border=THIN_BORDER,
            num_format=NUM_FMT2 if col in (1, 2, 4, 6, 7) and isinstance(v, (int, float)) else None,
        )

    # 일별
    ws['A7'] = '일별 매매기준율'
    _style(ws['A7'], font=FONT_BOLD)

    headers9 = ['날짜', '통화명', '환율', '전일대비', 'Cross Rate']
    for col, h in enumerate(headers9, 1):
        c = ws.cell(row=9, column=col, value=h)
        _style(c, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)

    daily_rows = list(rate_data.get('daily', []))
    display_start = str(rate_data.get('display_start', '') or '')
    display_end = str(rate_data.get('display_end', '') or '')
    if display_start:
        daily_rows = [d for d in daily_rows if str(d.get('date', '')) >= display_start]
    if display_end:
        daily_rows = [d for d in daily_rows if str(d.get('date', '')) <= display_end]

    for r, d in enumerate(daily_rows, 10):
        vals = [d['date'], rate_data.get('currency_name', ''), d['rate'], d['change'], d['cross']]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            _style(
                c, font=FONT_DEFAULT, align=CENTER if col != 2 else LEFT, border=THIN_BORDER,
                num_format=NUM_FMT2 if col in (3, 4) else None,
            )


# ── 쇼피 소포수령증 시트 작성 ───────────────────────────────────

def write_shopee_sheet(ws, shopee_data: dict, rates: dict, submitter: dict = None) -> int:
    """
    쇼피(MYR) 등 국가별 쇼피 시트 작성
    환율: 소포수령증 발행일(write_date) 기준
    """
    col_widths = [16, 3, 3, 12, 3, 3, 20, 3, 3, 3, 5, 3, 5, 3, 3, 12, 3, 10, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    currency   = shopee_data.get('currency', '')
    carrier    = shopee_data.get('carrier', '주)두라로지스틱스')
    country    = shopee_data.get('country', '')
    period_end = shopee_data.get('period_end', '')
    divisor    = RATE_DIVISOR.get(currency, 1)   # VND·JPY → 100, 나머지 → 1

    # ── 행 1: 제목 헤더 ──
    sub = submitter or shopee_data.get('submitter') or DEFAULT_SUBMITTER
    ws.merge_cells('A1:J1')
    ws['A1'] = (
        f"해외배송 소포 수령증\n"
        f"Registration No. 117-81-45551\n"
        f"{sub.get('biz_no','')}\n{sub.get('ceo','')}"
    )
    _style(ws['A1'], font=FONT_BOLD, align=CENTER)
    ws.row_dimensions[1].height = 55

    ws.merge_cells('L1:S1')
    ws['L1'] = (
        f"해외배송기간: {shopee_data.get('period_start','')} ~ {period_end}\n"
        f"{sub.get('name','')}\n{sub.get('address','')}"
    )
    _style(ws['L1'], font=FONT_DEFAULT, align=LEFT)

    # ── 행 2-4: 인적사항 라벨 ──
    ws['A2'] = '사업자등록번호\n대표자 성명 거래기간'
    ws['A3'] = '상호(법인명) 작성일자'
    ws['A4'] = '사업장소재지'
    for row in [2, 3, 4]:
        _style(ws.cell(row=row, column=1), font=FONT_DEFAULT, align=LEFT)

    # ── 행 5: 거래기간, 작성일자 ──
    ws['A5'] = f"{shopee_data.get('period_start','')} ~ {period_end}"
    ws['I5'] = shopee_data.get('write_date', '')
    _style(ws['A5'], font=FONT_DEFAULT)
    _style(ws['I5'], font=FONT_DEFAULT)

    # ── 행 6: 섹션 2 제목 ──
    ws.merge_cells('A6:S6')
    ws['A6'] = '2. 해외배송 소포 수령 수량'
    _style(ws['A6'], font=FONT_BOLD, fill=SUBHEAD_FILL)

    # ── 행 7: 헤더 ──
    headers7 = {
        'A': '해외배송업체', 'D': '배송국가', 'G': '기간', 'K': '통화', 'M': '발송수량', 'P': '발송금액'
    }
    for col_letter, val in headers7.items():
        c = ws[f'{col_letter}7']
        c.value = val
        _style(c, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)
    _merge_row(ws, 7, _RECEIPT_GROUPS_2, border=THIN_BORDER)
    for col_letter in headers7:
        _style(ws[f'{col_letter}7'], font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)

    # ── 행 8: 데이터 ──
    ws['A8'] = carrier
    ws['D8'] = country
    ws['G8'] = f"{shopee_data.get('period_start','')} ~ {period_end}"
    ws['K8'] = currency
    ws['M8'] = shopee_data.get('total_qty', 0)
    ws['P8'] = shopee_data.get('total_amount', 0.0)
    _merge_row(ws, 8, _RECEIPT_GROUPS_2, border=THIN_BORDER)
    for col in ['A', 'D', 'G', 'K', 'M', 'P']:
        nf = NUM_FMT if col == 'M' else (NUM_FMT2 if col == 'P' else None)
        _style(ws[f'{col}8'], font=FONT_DEFAULT, align=CENTER, border=THIN_BORDER, num_format=nf)

    # ── 행 9: 합계 ──
    ws['M9'] = shopee_data.get('total_qty', 0)
    ws['P9'] = shopee_data.get('total_amount', 0.0)
    ws['G9'] = '합계'
    _merge_row(ws, 9, _RECEIPT_GROUPS_2, border=THIN_BORDER)
    _style(ws['G9'], font=FONT_BOLD, align=CENTER, border=THIN_BORDER)
    _style(ws['M9'], font=FONT_BOLD, align=CENTER, border=THIN_BORDER, num_format=NUM_FMT)
    _style(ws['P9'], font=FONT_BOLD, align=CENTER, border=THIN_BORDER, num_format=NUM_FMT2)

    # ── 행 10: 섹션 3 제목 ──
    ws.merge_cells('A10:O10')
    ws['A10'] = '3. 해외배송 내역'
    _style(ws['A10'], font=FONT_BOLD, fill=SUBHEAD_FILL)

    # ── 행 11: 컬럼 헤더 ──
    col_headers = {
        'A': '해외배송업체', 'D': '발행일', 'G': '운송장번호',
        'K': '도착국가', 'M': '발송수량', 'P': '수출신고금액', 'R': '환율', 'S': '원화'
    }
    for col_letter, val in col_headers.items():
        c = ws[f'{col_letter}11']
        c.value = val
        _style(c, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)
    _merge_row(ws, 11, _RECEIPT_GROUPS_3, border=THIN_BORDER)
    for col_letter in col_headers:
        _style(ws[f'{col_letter}11'], font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)

    # ── 행 12+: 거래 데이터 (각 행의 발행일 기준 환율 개별 적용) ──
    row = 12
    total_krw = 0
    for tx in shopee_data.get('transactions', []):
        # 해당 거래의 발행일(tx['date']) 기준 환율 개별 조회
        tx_rate = _get_rate(rates, currency, tx['date'])
        krw = round(tx['amount'] * tx_rate / divisor)
        total_krw += krw

        ws.cell(row=row, column=1,  value=tx['carrier'])
        ws.cell(row=row, column=4,  value=tx['date'])
        ws.cell(row=row, column=7,  value=tx['tracking_no'])
        ws.cell(row=row, column=11, value=tx['country'])
        ws.cell(row=row, column=13, value=tx['qty'])
        ws.cell(row=row, column=16, value=tx['amount'])
        ws.cell(row=row, column=18, value=tx_rate)
        ws.cell(row=row, column=19, value=krw)

        _merge_row(ws, row, _RECEIPT_GROUPS_3, border=THIN_BORDER)
        for col in [1, 4, 7, 11, 13, 16, 18, 19]:
            c = ws.cell(row=row, column=col)
            nf = {13: NUM_FMT, 16: NUM_FMT2, 18: NUM_FMT2, 19: NUM_FMT}.get(col)
            _style(c, font=FONT_DEFAULT, align=CENTER if col != 1 else LEFT, border=THIN_BORDER, num_format=nf)

        row += 1

    # 원화 합계
    ws['S10'] = total_krw
    ws['Q10'] = shopee_data.get('total_amount', 0.0)

    # ── 푸터 ──
    footer_row = row + 1
    ws.merge_cells(f'A{footer_row}:S{footer_row}')
    ws[f'A{footer_row}'] = '상기 내역은 판매자가 두라로지스틱스를 통하여 해외 배송한 내역임을 증명합니다'
    _style(ws[f'A{footer_row}'], font=FONT_DEFAULT, align=CENTER)

    footer_row += 1
    ws[f'A{footer_row}'] = '상호 (법인명)'
    ws[f'C{footer_row}'] = '두라로지스틱스'
    ws[f'H{footer_row}'] = '사업자 등록번호'

    footer_row += 1
    ws[f'A{footer_row}'] = '사업장 소재지'
    ws[f'C{footer_row}'] = '서울특별시 강서구 금낭화로 54-7 (방화동, 동해빌딩 1층)'

    footer_row += 1
    ws[f'A{footer_row}'] = '비고'
    ws[f'C{footer_row}'] = '본 증명서를 위조하거나 변조하는 등 모든 행위에 대한 책임은 판매자에게 있습니다'

    footer_row += 1
    ws[f'A{footer_row}'] = '(주)두라로지스틱스'

    return total_krw


# ── 통화별 수출신고 템플릿 시트 작성 ─────────────────────────────

def write_currency_template_sheet(ws, currency: str,
                                   shopee_data: Optional[object],
                                   lazada_items: list,
                                   rates: dict,
                                   lazada_write_date: str = '',
                                   lazada_rate_override: float = None,
                                   ebay_items: list = None):
    """
    MYR, PHP, SGD 등 수출신고 프로그램용 시트 작성
    환율: 각 소포수령증 발행일(write_date) 기준
    """
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14

    # 라자다 환율: 큐텐과 동일하게 평균환율 사용
    lazada_rate = (lazada_rate_override if lazada_rate_override is not None
                   else rates.get(currency, {}).get('average', 0.0))
    divisor     = RATE_DIVISOR.get(currency, 1)   # VND·JPY → 100, 나머지 → 1

    # ── 쇼피 소계: 같은 통화의 PDF가 여러 개여도 모두 합산합니다. ──
    if isinstance(shopee_data, list):
        shopee_items = [sd for sd in shopee_data if sd]
    elif shopee_data:
        shopee_items = [shopee_data]
    else:
        shopee_items = []
    shopee_items = sorted(shopee_items, key=_shopee_sort_key)

    shopee_transactions = []
    shopee_fx  = 0.0
    shopee_krw = 0
    for sd in shopee_items:
        txs = sd.get('transactions', []) or []
        if txs:
            shopee_transactions.extend(txs)
            for tx in txs:
                tx_rate = _get_rate(rates, currency, tx.get('date', ''))
                amount = float(tx.get('amount', 0) or 0)
                shopee_fx  += amount
                shopee_krw += round(amount * tx_rate / divisor)
        else:
            # 거래별 상세가 없을 때는 소계만 합산합니다.
            amount = float(sd.get('total_amount', 0) or 0)
            rate_date = sd.get('write_date') or sd.get('period_end') or ''
            tx_rate = _get_rate(rates, currency, rate_date)
            shopee_fx += amount
            shopee_krw += round(amount * tx_rate / divisor)

    shopee_transactions.sort(
        key=lambda tx: (_date_to_int(tx.get('date')) or 0, str(tx.get('tracking_no', '')))
    )

    # ── 라자다 소계 ──
    lazada_fx  = sum(it['amount'] for it in lazada_items)
    lazada_krw = round(lazada_fx * lazada_rate / divisor)

    # ── 이베이/린코스 소계: 발행월 기준 월평균 환율 사용 ──
    from .exchange_rate import monthly_avg_rate_for_month
    ebay_items = ebay_items or []
    ebay_fx = 0.0
    ebay_krw = 0
    for it in ebay_items:
        amount = float(it.get('amount', 0) or 0)
        month_key = it.get('month') or it.get('date') or it.get('period_end') or ''
        rate = monthly_avg_rate_for_month(rates.get(currency), month_key)
        ebay_fx += amount
        ebay_krw += round(amount * rate / divisor)

    total_krw = shopee_krw + lazada_krw + ebay_krw

    # ── 행 1-3 요약 ──
    ws.cell(row=1, column=5, value='쇼피')
    ws.cell(row=1, column=6, value=shopee_fx)
    ws.cell(row=1, column=7, value=shopee_krw)
    ws.cell(row=2, column=5, value='라자다')
    ws.cell(row=2, column=6, value=lazada_fx)
    ws.cell(row=2, column=7, value=lazada_krw)
    ws.cell(row=3, column=5, value='이베이')
    ws.cell(row=3, column=6, value=ebay_fx)
    ws.cell(row=3, column=7, value=ebay_krw)

    for row in [1, 2, 3]:
        for col in [5, 6, 7]:
            c = ws.cell(row=row, column=col)
            nf = NUM_FMT2 if col == 6 else (NUM_FMT if col == 7 else None)
            _style(c, font=FONT_DEFAULT, align=RIGHT, num_format=nf)

    # ── 행 4: 헤더 ──
    headers = ['수출신고번호', '기타영세율건수', '선(기)적일자', '통화코드', '환율', '외화금액', '원화금액']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        _style(c, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)

    # ── 행 5+: 쇼피 거래 (여러 PDF를 한 시트에 합쳐 날짜순 정렬) ──
    data_row = 5
    for tx in shopee_transactions:
        tx_rate  = _get_rate(rates, currency, tx.get('date', ''))
        amount   = float(tx.get('amount', 0) or 0)
        krw      = round(amount * tx_rate / divisor)
        date_int = _date_to_int(tx.get('date', ''))
        row_vals = ['', 1, date_int, currency, tx_rate, amount, krw]
        for col, v in enumerate(row_vals, 1):
            c = ws.cell(row=data_row, column=col, value=v)
            nf = {5: NUM_FMT2, 6: NUM_FMT2, 7: NUM_FMT}.get(col)
            _style(c, font=FONT_DEFAULT, align=CENTER, border=THIN_BORDER, num_format=nf)
        data_row += 1

    # ── 라자다 거래 (발행일 기준 환율 적용) ──
    for it in lazada_items:
        krw = round(it['amount'] * lazada_rate / divisor)
        date_int_laz = _date_to_int(lazada_write_date)
        row_vals = ['', 1, date_int_laz, currency, lazada_rate, it['amount'], krw]
        for col, v in enumerate(row_vals, 1):
            c = ws.cell(row=data_row, column=col, value=v)
            nf = {5: NUM_FMT2, 6: NUM_FMT2, 7: NUM_FMT}.get(col)
            _style(c, font=FONT_DEFAULT, align=CENTER, border=THIN_BORDER, num_format=nf)
        data_row += 1

    # ── 이베이/린코스 거래 (발행월 기준 월평균 환율 적용) ──
    for it in ebay_items:
        month_key = it.get('month') or it.get('date') or it.get('period_end') or ''
        rate = monthly_avg_rate_for_month(rates.get(currency), month_key)
        amount = float(it.get('amount', 0) or 0)
        krw = round(amount * rate / divisor)
        date_int_ebay = _date_to_int(it.get('date') or it.get('period_end') or '')
        row_vals = ['', 1, date_int_ebay, currency, rate, amount, krw]
        for col, v in enumerate(row_vals, 1):
            c = ws.cell(row=data_row, column=col, value=v)
            nf = {5: NUM_FMT2, 6: NUM_FMT2, 7: NUM_FMT}.get(col)
            _style(c, font=FONT_DEFAULT, align=CENTER, border=THIN_BORDER, num_format=nf)
        data_row += 1


# ── 라자다 소포수령증 시트 ───────────────────────────────────────

def write_lazada_receipt_sheet(ws, lazada_data: dict, rates: dict, submitter: dict = None):
    """라자다(소포수령증) 시트"""
    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 15

    carrier    = lazada_data.get('carrier', '용성종합물류')
    period_end = lazada_data.get('period_end', '')

    # 헤더
    ws['A1'] = (
        f'YONG SUNG LOGISTICS CO., LTD.\n'
        f'ROOM 1215, TOWER A 152, MAGOKSEO-RO, GANGSEO-GU, SEOUL, KOREA\n'
        f'TEL: 82-2-2664-4032  FAX: 82-2-2664-3815\n'
        f'E-mail : admin@yslogic.co.kr    URL : http://www.yslogic.co.kr'
    )
    _style(ws['A1'], font=FONT_DEFAULT, align=LEFT)
    ws.row_dimensions[1].height = 55

    ws['A2'] = '해외화물 소포 수령증'
    _style(ws['A2'], font=FONT_TITLE, align=CENTER)

    ws['A3'] = '1.   제출자 인적 사항'
    _style(ws['A3'], font=FONT_BOLD)

    sub = submitter or lazada_data.get('submitter') or DEFAULT_SUBMITTER
    info_rows = [
        ('사업자등록번호', sub.get('biz_no', ''), '상호(법인명)', sub.get('name', '')),
        ('성명(대표자)',   sub.get('ceo', ''),    '사업장소재지', sub.get('address', '')),
        ('거래기간',
         f"{lazada_data.get('period_start','')} – {period_end}",
         '작성일자', lazada_data.get('write_date', '')),
    ]
    for r, (k1, v1, k2, v2) in enumerate(info_rows, 4):
        ws.cell(row=r, column=1, value=k1)
        ws.cell(row=r, column=4, value=v1)
        ws.cell(row=r, column=9, value=k2)
        ws.cell(row=r, column=11, value=v2)

    ws['A7'] = '2.   해외 배송 내역서'
    _style(ws['A7'], font=FONT_BOLD)
    ws['A8'] = '발행사유'
    ws['B8'] = f'{carrier}를 통해 해외로 수출한 내역 증명'

    # 헤더행
    header_row = 9
    for col, h in enumerate(['서비스', '해외배송업체', '출발', '도착', '발송번호', '발송수량', '금액'], 1):
        c = ws.cell(row=header_row, column=col, value=h)
        _style(c, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)

    # 데이터
    for r, it in enumerate(lazada_data.get('items', []), header_row + 1):
        row_vals = [
            it.get('service', '라자다'),
            it.get('carrier', carrier),
            it.get('origin', 'KR'),
            it.get('destination', ''),
            it.get('tracking_no', ''),
            f"{it.get('qty', '')}건",
            f"{it.get('amount', '')}({it.get('currency', '')})",
        ]
        for col, v in enumerate(row_vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            _style(c, font=FONT_DEFAULT, align=CENTER, border=THIN_BORDER)



# ── 이베이/린코스 소포수령증 시트 ───────────────────────────────

def write_ebay_receipt_sheet(ws, ebay_data: dict, rates: dict, submitter: dict = None):
    """이베이 - 린코스 해외배송 소포 수령증 시트"""
    from .exchange_rate import monthly_avg_rate_for_month

    for col, width in {'A': 16, 'B': 18, 'C': 18, 'D': 12, 'E': 10, 'F': 14, 'G': 12, 'H': 14}.items():
        ws.column_dimensions[col].width = width

    sub = submitter or ebay_data.get('submitter') or DEFAULT_SUBMITTER
    ws['A1'] = '해외배송 소포 수령증 - 이베이'
    _style(ws['A1'], font=FONT_TITLE, align=CENTER)
    ws.merge_cells('A1:H1')

    info = [
        ('사업자등록번호', sub.get('biz_no',''), '상호(법인명)', sub.get('name','')),
        ('성명(대표자)', sub.get('ceo',''), '사업장소재지', sub.get('address','')),
        ('거래기간', f"{ebay_data.get('period_start','')} ~ {ebay_data.get('period_end','')}", '작성일자', ebay_data.get('write_date','')),
    ]
    r = 3
    for k1, v1, k2, v2 in info:
        vals = [k1, v1, k2, v2]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            _style(cell, font=FONT_BOLD if c in [1,3] else FONT_DEFAULT, fill=HEADER_FILL if c in [1,3] else None, align=CENTER if c in [1,3] else LEFT, border=THIN_BORDER)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value='2. 해외배송 소포 수령증')
    _style(ws.cell(row=r, column=1), font=FONT_BOLD)
    r += 1
    headers = ['해외배송업체', '배송국가/Service', '현지송장번호', '통화단위', '발송수량', '신고금액']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=c, value=h)
        _style(cell, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)
    for item in ebay_data.get('summary_items', []):
        r += 1
        vals = [item.get('carrier',''), item.get('service',''), item.get('tracking_no',''), item.get('currency',''), item.get('qty',0), item.get('amount',0)]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            nf = NUM_FMT2 if c == 6 else (NUM_FMT if c == 5 else None)
            _style(cell, font=FONT_DEFAULT, align=CENTER if c != 6 else RIGHT, border=THIN_BORDER, num_format=nf)

    r += 2
    ws.cell(row=r, column=1, value='3. 해외배송 내역서')
    _style(ws.cell(row=r, column=1), font=FONT_BOLD)
    r += 1
    headers = ['발행월', '해외배송업체', '배송국가/Service', '통화단위', '발송수량', '신고금액', '월평균환율', '원화금액']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=c, value=h)
        _style(cell, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)
    for item in ebay_data.get('items', []):
        r += 1
        cur = item.get('currency','')
        rate = monthly_avg_rate_for_month(rates.get(cur), item.get('month') or '')
        div = RATE_DIVISOR.get(cur, 1)
        amount = float(item.get('amount',0) or 0)
        krw = round(amount * rate / div)
        vals = [item.get('month',''), item.get('carrier',''), item.get('service',''), cur, item.get('qty',0), amount, rate, krw]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            nf = {5: NUM_FMT, 6: NUM_FMT2, 7: NUM_FMT2, 8: NUM_FMT}.get(c)
            _style(cell, font=FONT_DEFAULT, align=CENTER if c not in [6,8] else RIGHT, border=THIN_BORDER, num_format=nf)

# ── 큐텐 소포수령증 시트 ────────────────────────────────────────

def write_qoo10_sheet(ws, qoo10_data: Optional[dict], jpy_rate: float, submitter: dict = None):
    """
    큐텐(소포수령증) 시트
    jpy_rate: 거래기간 마지막날 JPY 환율 (100엔 기준)
    """
    ws['A1'] = '해외배송 소포 수령증'
    _style(ws['A1'], font=FONT_TITLE, align=CENTER)

    ws['A3'] = '1.제출자 인적사항'
    _style(ws['A3'], font=FONT_BOLD)

    sub = submitter or DEFAULT_SUBMITTER
    ws['A5'] = '사업자등록번호'; ws['B5'] = sub.get('biz_no', '')
    ws['C5'] = '상호（법인명）'; ws['D5'] = sub.get('name', '')
    ws['A6'] = '성명 （대표자）'; ws['B6'] = sub.get('ceo', '')
    ws['C6'] = '사업장소재지'; ws['D6'] = sub.get('address', '')
    ws['A7'] = '거래기간'

    if qoo10_data:
        period = f"{qoo10_data.get('period_start','')} ~ {qoo10_data.get('period_end','')}"
        ws['B7'] = period

    ws['A9'] = '2.해외배송 소포 수령증'
    _style(ws['A9'], font=FONT_BOLD)

    headers = ['판매처', '해외배송업체', '배송국가', '송장번호', '수량', '비고']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=11, column=col, value=h)
        _style(c, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)

    if qoo10_data:
        ws.cell(row=12, column=1, value='Qoo10')
        ws.cell(row=12, column=2, value='국제로지스틱')
        ws.cell(row=12, column=3, value='일본')
        ws.cell(row=12, column=4, value=qoo10_data.get('tracking_no', ''))
        ws.cell(row=12, column=5, value=f"{qoo10_data.get('qty', '')} 건")

        ws.cell(row=13, column=1, value='당기 해외배송 합계')
        ws.cell(row=13, column=5, value=f"{qoo10_data.get('qty', '')} 건")

        ws['A15'] = '3. 해외배송 내역서'
        _style(ws['A15'], font=FONT_BOLD)
        ws['A17'] = '발행사유'; ws['B17'] = '국제로지스틱을 통해 해외로 수출한 내역 증명'

        detail_headers = ['판매처', '해외배송업체', '출발', '도착', '발송번호', '발송수량', '금액 (JPY)', '원화금액']
        for col, h in enumerate(detail_headers, 1):
            c = ws.cell(row=18, column=col, value=h)
            _style(c, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)

        entries = qoo10_data.get('entries') or [{
            'tracking_no': qoo10_data.get('tracking_no', ''),
            'qty':    qoo10_data.get('qty', 0),
            'amount': qoo10_data.get('amount', 0),
            'rate':   jpy_rate,
            'krw':    round(qoo10_data.get('amount', 0) * jpy_rate / 100),
        }]

        ws.cell(row=18, column=8, value='적용: 발행일별 환율 (100엔)')

        r = 19
        total_jpy = 0
        total_krw = 0
        total_qty = 0
        for e in entries:
            e_rate = e.get('rate', jpy_rate)
            e_amt  = e.get('amount', 0)
            e_krw  = e.get('krw', round(e_amt * e_rate / 100))
            e_qty  = e.get('qty', 0)
            ws.cell(row=r, column=1, value='Qoo10')
            ws.cell(row=r, column=2, value='국제로지스틱')
            ws.cell(row=r, column=3, value='KR')
            ws.cell(row=r, column=4, value='JP')
            ws.cell(row=r, column=5, value=e.get('tracking_no', ''))
            ws.cell(row=r, column=6, value=f"{e_qty} 건")
            ws.cell(row=r, column=7, value=e_amt)
            ws.cell(row=r, column=8, value=e_krw)
            for col in range(1, 9):
                nf = NUM_FMT if col in (7, 8) else None
                _style(ws.cell(row=r, column=col), font=FONT_DEFAULT, align=CENTER, border=THIN_BORDER, num_format=nf)
            total_jpy += e_amt
            total_krw += e_krw
            total_qty += e_qty
            r += 1

        ws.cell(row=r, column=1, value='당기 해외배송 합계')
        ws.cell(row=r, column=6, value=f"{total_qty} 건")
        ws.cell(row=r, column=7, value=total_jpy)
        ws.cell(row=r, column=8, value=total_krw)
        for col in range(1, 9):
            nf = NUM_FMT if col in (7, 8) else None
            _style(ws.cell(row=r, column=col), font=FONT_BOLD, align=CENTER, border=THIN_BORDER, num_format=nf)
    else:
        ws['A12'] = '⚠️ 큐텐 데이터 없음 — STEP 2에서 수동 입력하세요'
        _style(ws['A12'], font=Font(name='맑은 고딕', size=9, color='FF0000'))


# ── 총집계 시트 ─────────────────────────────────────────────────

def write_summary_sheet(ws, shopee_totals: dict, lazada_totals: dict,
                         qoo10_data: Optional[dict], jpy_rate: float,
                         year_month: str, submitter: dict = None,
                         ebay_totals: dict = None):
    """총집계 시트 작성.

    실제 집계된 플랫폼만 표시합니다. 예를 들어 이베이 자료만 처리한 경우
    쇼피/라자다/큐텐 구역은 만들지 않습니다. 각 플랫폼 안에서도 실제로
    집계된 통화만 행으로 출력합니다.
    """
    NUM  = '#,##0'
    NUM2 = '#,##0.00'
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 16

    sub = submitter or DEFAULT_SUBMITTER
    ws['B1'] = f"{sub.get('name','')}({sub.get('biz_no','')})"
    _style(ws['B1'], font=FONT_TITLE)
    ws['D2'] = year_month
    _style(ws['D2'], font=FONT_BOLD)

    def _sub(r, label):
        c = ws.cell(row=r, column=2, value=label)
        _style(c, font=FONT_BOLD, fill=SUBHEAD_FILL, align=CENTER)

    def _hdr3(r, h1, h2, h3):
        for col, val in [(2, h1), (3, h2), (4, h3)]:
            c = ws.cell(row=r, column=col, value=val)
            _style(c, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)

    def _datarow(r, name, fx, krw):
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=fx)
        ws.cell(row=r, column=4, value=krw)
        _style(ws.cell(row=r, column=2), font=FONT_DEFAULT, align=LEFT,  border=THIN_BORDER)
        _style(ws.cell(row=r, column=3), font=FONT_DEFAULT, align=RIGHT, border=THIN_BORDER, num_format=NUM2)
        _style(ws.cell(row=r, column=4), font=FONT_DEFAULT, align=RIGHT, border=THIN_BORDER, num_format=NUM)

    def _totalrow(r, krw):
        ws.cell(row=r, column=2, value='총합')
        ws.cell(row=r, column=4, value=krw)
        _style(ws.cell(row=r, column=2), font=FONT_BOLD, align=LEFT,  border=THIN_BORDER, fill=GRAY_FILL)
        _style(ws.cell(row=r, column=3), border=THIN_BORDER, fill=GRAY_FILL)
        _style(ws.cell(row=r, column=4), font=FONT_BOLD, align=RIGHT, border=THIN_BORDER, fill=GRAY_FILL, num_format=NUM)

    COUNTRY_NAMES = {
        'MYR': '말레이시아(MYR)', 'PHP': '필리핀(PHP)',
        'SGD': '싱가폴(SGD)', 'THB': '태국(THB)',
        'TWD': '대만(TWD)', 'VND': '베트남(VND)',
        'BRL': '브라질(BRL)', 'MXN': '멕시코(MXN)',
        'USD': '미국(USD)', 'EUR': '유로(EUR)', 'GBP': '영국(GBP)',
        'CAD': '캐나다(CAD)', 'AUD': '호주(AUD)',
    }

    # 실제 집계된 통화만 대상으로 삼습니다. totals 딕셔너리는 실제 입력
    # 자료가 있을 때만 통화 키가 생성되므로 값이 0인 정상 거래도 유지됩니다.
    shopee_totals = shopee_totals or {}
    lazada_totals = lazada_totals or {}
    ebay_totals = ebay_totals or {}

    row = 5
    written_platforms = 0

    if shopee_totals:
        _sub(row, '쇼피')
        _hdr3(row + 1, '국가', '외화', '원화')
        data_row = row + 2
        total_krw = 0
        for cur in _ordered_currencies(shopee_totals.keys()):
            data = shopee_totals.get(cur, {})
            fx = data.get('fx', 0.0)
            krw = data.get('krw', 0)
            total_krw += krw
            _datarow(data_row, COUNTRY_NAMES.get(cur, cur), fx, krw)
            data_row += 1
        _totalrow(data_row, total_krw)
        row = data_row + 2
        written_platforms += 1

    if lazada_totals:
        _sub(row, '라자다')
        _hdr3(row + 1, '국가', '외화', '원화')
        data_row = row + 2
        total_krw = 0
        for cur in _ordered_currencies(lazada_totals.keys()):
            data = lazada_totals.get(cur, {})
            fx = data.get('fx', 0.0)
            krw = data.get('krw', 0)
            total_krw += krw
            _datarow(data_row, COUNTRY_NAMES.get(cur, cur), fx, krw)
            data_row += 1
        _totalrow(data_row, total_krw)
        row = data_row + 2
        written_platforms += 1

    if ebay_totals:
        _sub(row, '이베이')
        _hdr3(row + 1, '통화', '외화', '원화')
        data_row = row + 2
        total_krw = 0
        for cur in _ordered_currencies(ebay_totals.keys()):
            data = ebay_totals.get(cur, {})
            fx = data.get('fx', 0.0)
            krw = data.get('krw', 0)
            total_krw += krw
            _datarow(data_row, COUNTRY_NAMES.get(cur, cur), fx, krw)
            data_row += 1
        _totalrow(data_row, total_krw)
        row = data_row + 2
        written_platforms += 1

    if qoo10_data:
        _sub(row, '큐텐')
        _hdr3(row + 1, '외화', '평균환율', '원화')
        jpy_amount = qoo10_data.get('amount', 0)
        krw = qoo10_data.get('total_krw') or round(jpy_amount * jpy_rate / 100)
        # 평균환율 = 실효환율(원화÷외화×100) — 외화·원화와 정확히 일치
        eff_rate = round(krw * 100 / jpy_amount, 2) if jpy_amount else jpy_rate
        data_row = row + 2
        ws.cell(row=data_row, column=2, value=jpy_amount)
        ws.cell(row=data_row, column=3, value=eff_rate)
        ws.cell(row=data_row, column=4, value=krw)
        _style(ws.cell(row=data_row, column=2), font=FONT_DEFAULT, align=RIGHT, border=THIN_BORDER, num_format=NUM)
        _style(ws.cell(row=data_row, column=3), font=FONT_DEFAULT, align=RIGHT, border=THIN_BORDER, num_format=NUM2)
        _style(ws.cell(row=data_row, column=4), font=FONT_DEFAULT, align=RIGHT, border=THIN_BORDER, num_format=NUM)
        written_platforms += 1

    # 방어 로직: 데이터 없는 상태에서도 워크북 구조가 깨지지 않게 안내만 표시합니다.
    if written_platforms == 0:
        ws['B5'] = '집계된 플랫폼이 없습니다.'
        _style(ws['B5'], font=FONT_DEFAULT)



# ── 사용 데이터 기준 시트 정리 유틸 ───────────────────────────────
PREFERRED_CURRENCY_ORDER = ['MYR', 'PHP', 'SGD', 'THB', 'TWD', 'VND', 'BRL', 'MXN', 'USD', 'EUR', 'GBP', 'CAD', 'AUD', 'JPY']
SHOPEE_SHEET_NAMES = {
    'MYR': '쇼피(MYR)', 'PHP': '쇼피(PHP)', 'SGD': '쇼피(SGD)',
    'THB': '쇼피(THB)', 'TWD': '쇼피(TWD)', 'VND': '쇼피(VND)',
    'BRL': '쇼피(BRL)', 'MXN': '쇼피(MXN)',
}
LAZADA_CURRENCY_ORDER = ['MYR', 'PHP', 'SGD', 'VND']


def _ordered_currencies(values):
    """PREFERRED_CURRENCY_ORDER 기준으로 통화코드를 정렬합니다."""
    values = {str(v or '').upper() for v in values if v}
    return [cur for cur in PREFERRED_CURRENCY_ORDER if cur in values] + sorted(values - set(PREFERRED_CURRENCY_ORDER))


def _shopee_sort_key(sd):
    """같은 통화의 쇼피 PDF가 여러 개일 때 기간순으로 정렬합니다."""
    return (
        _date_to_int(sd.get('period_start')) or 0,
        _date_to_int(sd.get('period_end')) or 0,
        _date_to_int(sd.get('write_date')) or 0,
    )


def _shopee_items_for_currency(shopee_results, currency):
    """해당 통화의 쇼피 PDF 결과를 기간순으로 반환합니다."""
    return sorted(
        [s for s in (shopee_results or [])
         if s.get('currency') == currency and _has_shopee_data(s)],
        key=_shopee_sort_key,
    )


def _base_shopee_sheet_name(currency):
    return SHOPEE_SHEET_NAMES.get(currency, f'쇼피({currency})')


def _shopee_sheet_names_for_results(shopee_results, shopee_currencies):
    """
    쇼피 원본 PDF 시트명을 생성합니다.
    같은 통화의 PDF가 여러 개면 쇼피(MYR)(1), 쇼피(MYR)(2)처럼 번호를 붙입니다.
    한 개뿐이면 기존처럼 쇼피(MYR)를 유지합니다.
    """
    out = []
    for cur in shopee_currencies:
        items = _shopee_items_for_currency(shopee_results, cur)
        base = _base_shopee_sheet_name(cur)
        if len(items) <= 1:
            if items:
                out.append(base)
        else:
            out.extend(f'{base}({idx})' for idx, _sd in enumerate(items, 1))
    return out


def _has_shopee_data(sd):
    if not sd:
        return False
    if sd.get('transactions'):
        return True
    return bool(sd.get('currency') and (sd.get('total_qty', 0) or sd.get('total_amount', 0)))


def _has_lazada_data(lazada_result):
    return bool(lazada_result and lazada_result.get('items'))


def _has_ebay_data(ebay_result):
    return bool(ebay_result and ebay_result.get('items'))


def _ebay_sheet_names_for_results(ebay_results):
    items = [er for er in (ebay_results or []) if _has_ebay_data(er)]
    if not items:
        return []
    if len(items) == 1:
        return ['이베이']
    return [f'이베이({i})' for i, _ in enumerate(items, 1)]


def _has_qoo10_data(qoo10_result):
    if not qoo10_result:
        return False
    if qoo10_result.get('entries'):
        return True
    return bool(qoo10_result.get('qty', 0) or qoo10_result.get('amount', 0) or qoo10_result.get('tracking_no'))


def _infer_used_sources_and_currencies(shopee_results, lazada_result, qoo10_result, ebay_results=None):
    """
    실제 입력 데이터가 있는 소스/통화만 추려냅니다.
    이 결과를 기준으로 불필요한 쇼피/라자다/큐텐/이베이/환율/통화시트를 만들지 않습니다.
    """
    shopee_currencies = {sd.get('currency') for sd in (shopee_results or []) if _has_shopee_data(sd)}
    lazada_currencies = set()
    if _has_lazada_data(lazada_result):
        lazada_currencies = {it.get('currency') for it in lazada_result.get('items', []) if it.get('currency')}
    ebay_currencies = set()
    for er in (ebay_results or []):
        if _has_ebay_data(er):
            ebay_currencies.update(it.get('currency') for it in er.get('items', []) if it.get('currency'))
    qoo10_used = _has_qoo10_data(qoo10_result)
    used_currencies = set(shopee_currencies) | set(lazada_currencies) | set(ebay_currencies)
    if qoo10_used:
        used_currencies.add('JPY')
    return {
        'shopee_currencies': _ordered_currencies(shopee_currencies),
        'lazada_currencies': _ordered_currencies(lazada_currencies),
        'ebay_currencies': _ordered_currencies(ebay_currencies),
        'qoo10_used': qoo10_used,
        'used_currencies': _ordered_currencies(used_currencies),
    }


def _prune_workbook_sheets(wb, keep_sheet_names):
    """혹시 생성 과정에서 남은 불필요 시트를 최종적으로 삭제합니다."""
    keep = set(keep_sheet_names)
    for sheet_name in list(wb.sheetnames):
        if sheet_name not in keep:
            del wb[sheet_name]

# ── 월별집계 시트 ───────────────────────────────────────────────

def _month_label_from_date(value):
    """날짜 문자열/숫자에서 'YYYY년 MM월' 라벨을 반환합니다."""
    d = re.sub(r"\D", "", str(value or ""))[:8]
    if len(d) >= 6:
        return f"{d[:4]}년 {d[4:6]}월"
    return "날짜미상"


def write_monthly_summary_sheet(ws, shopee_results: list, lazada_result: Optional[dict],
                                qoo10_result: Optional[dict], rates: dict,
                                lazada_avg_rates: dict = None,
                                lazada_write_date: str = '',
                                jpy_rate: float = 0.0,
                                submitter: dict = None,
                                ebay_results: list = None):
    """
    월별집계 시트 작성.
    기준일은 각 문서의 수출실적/통화 시트에 들어가는 선(기)적일자와 동일하게 봅니다.
    - 쇼피: 거래별 발행일(tx['date'])
    - 라자다: 라자다 작성일자(write_date) 또는 기간 종료일
    - 큐텐: 입력 건별 발행일(write_date)
    """
    NUM = '#,##0'
    NUM2 = '#,##0.00'
    lazada_avg_rates = lazada_avg_rates or {}

    for col, width in {'A': 13, 'B': 12, 'C': 10, 'D': 10, 'E': 16, 'F': 16}.items():
        ws.column_dimensions[col].width = width

    sub = submitter or DEFAULT_SUBMITTER
    ws['A1'] = f"월별집계 - {sub.get('name','')}({sub.get('biz_no','')})"
    _style(ws['A1'], font=FONT_TITLE)
    ws.merge_cells('A1:F1')

    headers = ['월', '구분', '통화코드', '건수', '외화금액', '원화금액']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        _style(c, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)

    rows = []

    # 쇼피: 거래별 발행일 기준
    for sd in shopee_results or []:
        cur = sd.get('currency', '')
        if not cur:
            continue
        div = RATE_DIVISOR.get(cur, 1)
        txs = sd.get('transactions') or []
        if txs:
            for tx in txs:
                amount = float(tx.get('amount', 0) or 0)
                qty = int(tx.get('qty', 0) or 0)
                rate = _get_rate(rates, cur, tx.get('date', ''))
                krw = round(amount * rate / div)
                rows.append({
                    'month': _month_label_from_date(tx.get('date', '')),
                    'source': '쇼피', 'currency': cur, 'qty': qty,
                    'fx': amount, 'krw': krw,
                })
        else:
            amount = float(sd.get('total_amount', 0) or 0)
            qty = int(sd.get('total_qty', 0) or 0)
            rate_date = sd.get('write_date') or sd.get('period_end') or ''
            rate = _get_rate(rates, cur, rate_date)
            krw = round(amount * rate / div)
            rows.append({
                'month': _month_label_from_date(rate_date),
                'source': '쇼피', 'currency': cur, 'qty': qty,
                'fx': amount, 'krw': krw,
            })

    # 라자다: 작성일자/기간종료일 기준
    if lazada_result and lazada_result.get('items'):
        date_value = lazada_write_date or lazada_result.get('write_date') or lazada_result.get('period_end') or ''
        for it in lazada_result.get('items', []):
            cur = it.get('currency', '')
            if not cur:
                continue
            amount = float(it.get('amount', 0) or 0)
            qty = int(it.get('qty', 0) or 0)
            div = RATE_DIVISOR.get(cur, 1)
            rate = lazada_avg_rates.get(cur, rates.get(cur, {}).get('average', 0.0))
            krw = round(amount * rate / div)
            rows.append({
                'month': _month_label_from_date(date_value),
                'source': '라자다', 'currency': cur, 'qty': qty,
                'fx': amount, 'krw': krw,
            })

    # 이베이/린코스: 발행월 기준 월평균 환율
    from .exchange_rate import monthly_avg_rate_for_month
    for er in ebay_results or []:
        for it in er.get('items', []):
            cur = it.get('currency', '')
            if not cur:
                continue
            amount = float(it.get('amount', 0) or 0)
            qty = int(it.get('qty', 0) or 0)
            div = RATE_DIVISOR.get(cur, 1)
            rate = monthly_avg_rate_for_month(rates.get(cur), it.get('month') or it.get('date') or '')
            krw = round(amount * rate / div)
            rows.append({
                'month': _month_label_from_date((it.get('month') or '').replace('-', '') + '01'),
                'source': '이베이', 'currency': cur, 'qty': qty,
                'fx': amount, 'krw': krw,
            })

    # 큐텐: 입력 건별 발행일 기준
    if qoo10_result and qoo10_result.get('entries'):
        for e in qoo10_result.get('entries', []):
            amount = float(e.get('amount', 0) or 0)
            qty = int(e.get('qty', 0) or 0)
            rate = e.get('rate', jpy_rate)
            krw = e.get('krw', round(amount * rate / 100))
            date_value = e.get('write_date') or qoo10_result.get('write_date') or qoo10_result.get('period_end') or ''
            rows.append({
                'month': _month_label_from_date(date_value),
                'source': '큐텐', 'currency': 'JPY', 'qty': qty,
                'fx': amount, 'krw': krw,
            })

    # 집계: 월 + 구분 + 통화
    grouped = {}
    for item in rows:
        key = (item['month'], item['source'], item['currency'])
        if key not in grouped:
            grouped[key] = {'qty': 0, 'fx': 0.0, 'krw': 0}
        grouped[key]['qty'] += item['qty']
        grouped[key]['fx'] += item['fx']
        grouped[key]['krw'] += item['krw']

    # 출력
    r = 4
    grand_qty = 0
    grand_fx_by_currency = {}
    grand_krw = 0
    months = sorted({k[0] for k in grouped.keys()})
    source_order = {'쇼피': 1, '라자다': 2, '이베이': 3, '큐텐': 4}

    for month in months:
        month_qty = 0
        month_krw = 0
        month_fx_by_currency = {}
        keys = sorted(
            [k for k in grouped.keys() if k[0] == month],
            key=lambda x: (source_order.get(x[1], 99), PREFERRED_CURRENCY_ORDER.index(x[2]) if x[2] in PREFERRED_CURRENCY_ORDER else 99, x[2])
        )
        for _month, source, cur in keys:
            data = grouped[(_month, source, cur)]
            vals = [month, source, cur, data['qty'], data['fx'], data['krw']]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=r, column=col, value=v)
                nf = {4: NUM, 5: NUM2, 6: NUM}.get(col)
                _style(c, font=FONT_DEFAULT, align=CENTER if col <= 4 else RIGHT, border=THIN_BORDER, num_format=nf)
            month_qty += data['qty']
            month_fx_by_currency[cur] = month_fx_by_currency.get(cur, 0.0) + data['fx']
            month_krw += data['krw']
            grand_qty += data['qty']
            grand_fx_by_currency[cur] = grand_fx_by_currency.get(cur, 0.0) + data['fx']
            grand_krw += data['krw']
            r += 1

        # 월 합계: 외화는 통화가 여러 개일 수 있으므로 원화 합계 중심으로 표시
        ws.cell(row=r, column=1, value=month)
        ws.cell(row=r, column=2, value='월 합계')
        ws.cell(row=r, column=4, value=month_qty)
        ws.cell(row=r, column=6, value=month_krw)
        for col in range(1, 7):
            c = ws.cell(row=r, column=col)
            nf = {4: NUM, 6: NUM}.get(col)
            _style(c, font=FONT_BOLD, fill=GRAY_FILL, align=CENTER if col <= 4 else RIGHT, border=THIN_BORDER, num_format=nf)
        r += 1

    # 전체 총합
    ws.cell(row=r, column=1, value='전체')
    ws.cell(row=r, column=2, value='전체 총합')
    ws.cell(row=r, column=4, value=grand_qty)
    ws.cell(row=r, column=6, value=grand_krw)
    for col in range(1, 7):
        c = ws.cell(row=r, column=col)
        nf = {4: NUM, 6: NUM}.get(col)
        _style(c, font=FONT_BOLD, fill=SUBHEAD_FILL, align=CENTER if col <= 4 else RIGHT, border=THIN_BORDER, num_format=nf)

    # 참고: 통화별 전체 외화 합계
    r += 2
    ws.cell(row=r, column=1, value='통화별 외화 합계')
    _style(ws.cell(row=r, column=1), font=FONT_BOLD)
    r += 1
    for cur in _ordered_currencies(grand_fx_by_currency.keys()):
        ws.cell(row=r, column=2, value=cur)
        ws.cell(row=r, column=5, value=grand_fx_by_currency.get(cur, 0.0))
        _style(ws.cell(row=r, column=2), font=FONT_DEFAULT, align=CENTER, border=THIN_BORDER)
        _style(ws.cell(row=r, column=5), font=FONT_DEFAULT, align=RIGHT, border=THIN_BORDER, num_format=NUM2)
        r += 1

# ── 전체 엑셀 생성 ───────────────────────────────────────────────

def generate_excel(
    shopee_results: list,      # [parse_shopee_pdf() 결과, ...]
    lazada_result:  Optional[dict],   # parse_lazada_pdf() 결과
    qoo10_result:   Optional[dict],   # parse_qoo10_pdf() 결과
    rates:          dict,      # fetch_all_currencies() 결과
    output_path:    str,
    year:           int,
    month:          int,
    ebay_results:   Optional[list] = None,
):
    """전체 엑셀 파일 생성"""
    wb = Workbook()
    wb.remove(wb.active)

    ebay_results = ebay_results or []
    usage = _infer_used_sources_and_currencies(shopee_results, lazada_result, qoo10_result, ebay_results)
    shopee_currencies = usage['shopee_currencies']
    lazada_currencies = usage['lazada_currencies']
    qoo10_used = usage['qoo10_used']
    ebay_currencies = usage.get('ebay_currencies', [])
    used_currencies = usage['used_currencies']

    # ── 라자다 발행일 추출 (write_date → period_end fallback) ──
    if lazada_result:
        lazada_write_date = (lazada_result.get('write_date', '')
                             or lazada_result.get('period_end', ''))
    else:
        lazada_write_date = ''

    from .exchange_rate import avg_rate_for_period, monthly_avg_rate_for_month

    # ── 라자다 거래기간 평균환율 (통화별) ──
    lazada_avg_rates = {}
    if lazada_result:
        _lp_s = lazada_result.get('period_start', '')
        _lp_e = lazada_result.get('period_end', '')
        for _it in lazada_result.get('items', []):
            _lc = _it.get('currency', '')
            if _lc and _lc not in lazada_avg_rates:
                lazada_avg_rates[_lc] = avg_rate_for_period(rates.get(_lc), _lp_s, _lp_e)

    # ── 큐텐 JPY 환율: 거래기간 평균환율 사용 ──
    jpy_rate_data = rates.get('JPY')
    if jpy_rate_data:
        jpy_rate = jpy_rate_data.get('average', 0.0)
        if jpy_rate == 0.0:
            daily = jpy_rate_data.get('daily', [])
            if daily:
                jpy_rate = round(sum(d['rate'] for d in daily) / len(daily), 2)
    else:
        jpy_rate = 0.0
    # write_date 보존 (선적일자 기재용)
    qoo10_write_date = ''
    if qoo10_result:
        qoo10_write_date = (qoo10_result.get('write_date', '')
                            or qoo10_result.get('period_end', ''))
        # 큐텐 전체 거래기간 평균환율을 대표 환율로 사용 (표시·폴백용)
        _qs = qoo10_result.get('period_start', '')
        _qe = qoo10_result.get('period_end', '')
        if (_qs or _qe) and jpy_rate_data:
            _qavg = avg_rate_for_period(jpy_rate_data, _qs, _qe)
            if _qavg:
                jpy_rate = _qavg

        # ── 큐텐 건별 환율·원화 계산 (entries) ──
        q_entries = qoo10_result.get('entries')
        if not q_entries:
            q_entries = [{
                'tracking_no': qoo10_result.get('tracking_no', ''),
                'qty':         qoo10_result.get('qty', 0),
                'amount':      qoo10_result.get('amount', 0),
                'write_date':  qoo10_result.get('write_date', ''),
            }]
        q_total_krw = 0
        for e in q_entries:
            ps = e.get('period_start', '') or qoo10_result.get('period_start', '')
            pe = e.get('period_end', '') or qoo10_result.get('period_end', '')
            r = avg_rate_for_period(jpy_rate_data, ps, pe) or jpy_rate
            e['rate'] = r
            e['krw']  = round(e.get('amount', 0) * r / 100)
            q_total_krw += e['krw']
        qoo10_result['entries']   = q_entries
        qoo10_result['amount']    = sum(e.get('amount', 0) for e in q_entries)
        qoo10_result['qty']       = sum(e.get('qty', 0) for e in q_entries)
        qoo10_result['total_krw'] = q_total_krw

    # ── 제출자(판매자) 정보: PDF에서 자동 추출, 없으면 기본값 ──
    report_submitter = None
    for sd in shopee_results:
        if sd.get('submitter') and sd['submitter'].get('name'):
            report_submitter = sd['submitter']
            break
    if report_submitter is None and lazada_result and lazada_result.get('submitter', {}).get('name'):
        report_submitter = lazada_result['submitter']
    if report_submitter is None:
        for er in ebay_results:
            if er.get('submitter') and er['submitter'].get('name'):
                report_submitter = er['submitter']
                break
    if report_submitter is None and qoo10_result and qoo10_result.get('submitter', {}).get('name'):
        report_submitter = qoo10_result['submitter']
    if report_submitter is None:
        report_submitter = DEFAULT_SUBMITTER

    # ── 총집계 ──────────────────────────────────────────────
    ws_summary = wb.create_sheet('총집계')
    shopee_totals = {}
    lazada_totals = {}
    ebay_totals = {}

    for sd in shopee_results:
        cur = sd.get('currency', '')
        if not cur:
            continue
        # 각 거래의 발행일(tx['date']) 기준 환율로 개별 계산 후 합산
        div = RATE_DIVISOR.get(cur, 1)
        total_fx  = 0.0
        total_krw = 0
        for tx in sd.get('transactions', []):
            tx_rate    = _get_rate(rates, cur, tx['date'])
            total_fx  += tx['amount']
            total_krw += round(tx['amount'] * tx_rate / div)
        # 거래 내역 없으면 total_amount 사용 (fallback)
        if not sd.get('transactions'):
            rate_date = sd.get('write_date', '') or sd.get('period_end', '')
            rate = _get_rate(rates, cur, rate_date)
            total_fx  = sd.get('total_amount', 0.0)
            total_krw = round(total_fx * rate / div)
        if cur not in shopee_totals:
            shopee_totals[cur] = {'fx': 0.0, 'krw': 0}
        shopee_totals[cur]['fx'] += total_fx
        shopee_totals[cur]['krw'] += total_krw

    if lazada_result:
        laz_rate_by_cur = {}
        for it in lazada_result.get('items', []):
            cur = it.get('currency', '')
            if cur not in laz_rate_by_cur:
                laz_rate_by_cur[cur] = lazada_avg_rates.get(cur, rates.get(cur, {}).get('average', 0.0))
            rate = laz_rate_by_cur[cur]
            div  = RATE_DIVISOR.get(cur, 1)
            krw  = round(it.get('amount', 0.0) * rate / div)
            if cur not in lazada_totals:
                lazada_totals[cur] = {'fx': 0.0, 'krw': 0}
            lazada_totals[cur]['fx']  += it.get('amount', 0.0)
            lazada_totals[cur]['krw'] += krw

    # 이베이/린코스: 발행월 기준 월평균 매매기준율 사용
    for er in ebay_results:
        for it in er.get('items', []):
            cur = it.get('currency', '')
            if not cur:
                continue
            rate = monthly_avg_rate_for_month(rates.get(cur), it.get('month') or it.get('date') or '')
            div = RATE_DIVISOR.get(cur, 1)
            amount = float(it.get('amount', 0) or 0)
            krw = round(amount * rate / div)
            if cur not in ebay_totals:
                ebay_totals[cur] = {'fx': 0.0, 'krw': 0}
            ebay_totals[cur]['fx'] += amount
            ebay_totals[cur]['krw'] += krw

    period_label, _ = period_labels(shopee_results, lazada_result, qoo10_result, ebay_results=ebay_results,
                                    fallback=f'{year}년 {month:02d}월')

    write_summary_sheet(ws_summary, shopee_totals, lazada_totals,
                        qoo10_result, jpy_rate,
                        period_label, submitter=report_submitter,
                        ebay_totals=ebay_totals)

    # ── 월별집계 ─────────────────────────────────────────────
    # 총집계 바로 오른쪽에 배치합니다.
    ws_monthly = wb.create_sheet('월별집계', 1)
    write_monthly_summary_sheet(ws_monthly, shopee_results, lazada_result, qoo10_result,
                                rates, lazada_avg_rates=lazada_avg_rates,
                                lazada_write_date=lazada_write_date,
                                jpy_rate=jpy_rate, submitter=report_submitter,
                                ebay_results=ebay_results)

    # ── 통화별 수출신고 템플릿 시트
    # 실제 쇼피/라자다/큐텐 데이터가 있는 통화만 생성합니다.
    for cur in used_currencies:
        if cur == 'JPY':
            continue
        ws = wb.create_sheet(cur)
        sd = _shopee_items_for_currency(shopee_results, cur)
        lazada_items = []
        if _has_lazada_data(lazada_result):
            lazada_items = [it for it in lazada_result.get('items', [])
                            if it.get('currency') == cur]
        ebay_items = []
        for er in ebay_results:
            ebay_items.extend([it for it in er.get('items', []) if it.get('currency') == cur])
        write_currency_template_sheet(ws, cur, sd, lazada_items, rates,
                                      lazada_write_date=lazada_write_date,
                                      lazada_rate_override=lazada_avg_rates.get(cur),
                                      ebay_items=ebay_items)

    # ── JPY 수출신고 시트 (큐텐 데이터가 있을 때만 생성) ──
    if qoo10_used:
        ws_jpy = wb.create_sheet('JPY')
        headers = ['수출신고번호', '기타영세율건수', '선(기)적일자', '통화코드', '환율', '외화금액', '원화금액']
        for col, h in enumerate(headers, 1):
            c = ws_jpy.cell(row=4, column=col, value=h)
            _style(c, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)
        ws_jpy.cell(row=1, column=5, value='큐텐')
        ws_jpy.cell(row=1, column=6, value=qoo10_result.get('amount', 0)).number_format = NUM_FMT
        ws_jpy.cell(row=1, column=7, value=qoo10_result.get('total_krw', 0)).number_format = NUM_FMT
        _jr = 5
        for e in qoo10_result.get('entries', []):
            wd = e.get('write_date', '') or qoo10_write_date
            date_str = ''
            if wd:
                try:
                    date_str = int(str(wd).replace('-', '').replace('.', ''))
                except ValueError:
                    date_str = ''
            tracking = e.get('tracking_no') or qoo10_result.get('tracking_no', '')
            ws_jpy.cell(row=_jr, column=1, value='')
            ws_jpy.cell(row=_jr, column=2, value=1)
            ws_jpy.cell(row=_jr, column=3, value=date_str or None)
            ws_jpy.cell(row=_jr, column=4, value='JPY')
            ws_jpy.cell(row=_jr, column=5, value=e.get('rate', jpy_rate)).number_format = NUM_FMT2
            ws_jpy.cell(row=_jr, column=6, value=e.get('amount', 0)).number_format = NUM_FMT
            ws_jpy.cell(row=_jr, column=7, value=e.get('krw', 0)).number_format = NUM_FMT
            _jr += 1

        # ── 큐텐(소포수령증) ──
        ws_q10 = wb.create_sheet('큐텐(소포수령증)')
        write_qoo10_sheet(ws_q10, qoo10_result, jpy_rate, submitter=report_submitter)

    # ── 쇼피 원본 PDF별 시트
    # 같은 통화의 PDF가 여러 개면 쇼피(MYR)(1), 쇼피(MYR)(2)처럼 기간순으로 분리합니다.
    for cur in shopee_currencies:
        items = _shopee_items_for_currency(shopee_results, cur)
        base_sheet_name = _base_shopee_sheet_name(cur)
        for idx, sd in enumerate(items, 1):
            sheet_name = f'{base_sheet_name}({idx})' if len(items) > 1 else base_sheet_name
            ws = wb.create_sheet(sheet_name)
            write_shopee_sheet(ws, sd, rates, submitter=sd.get('submitter') or report_submitter)

    # ── 라자다(소포수령증) + 라자다(국가별)
    # 라자다 PDF가 있을 때만 생성하고, 통화별 라자다 시트도 실제 통화만 생성합니다.
    if _has_lazada_data(lazada_result):
        ws_laz = wb.create_sheet('라자다(소포수령증)')
        write_lazada_receipt_sheet(ws_laz, lazada_result, rates, submitter=report_submitter)

        for cur in lazada_currencies:
            ws = wb.create_sheet(f'라자다({cur})')
            items = [it for it in lazada_result.get('items', []) if it.get('currency') == cur]
            if items:
                headers = ['No', 'OBD DT', 'HBL No', 'MBL No', 'POL', 'POD', 'PKG', 'PKG Unit', 'G.WT', 'C.WT']
                for col, h in enumerate(headers, 1):
                    c = ws.cell(row=2, column=col, value=h)
                    _style(c, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)

    # ── 이베이/린코스 원본 PDF별 시트
    ebay_sheet_names = _ebay_sheet_names_for_results(ebay_results)
    for idx, er in enumerate([x for x in ebay_results if _has_ebay_data(x)], 0):
        ws_ebay = wb.create_sheet(ebay_sheet_names[idx])
        write_ebay_receipt_sheet(ws_ebay, er, rates, submitter=er.get('submitter') or report_submitter)

    # ── 환율 시트
    # 실제 데이터가 있는 통화만 생성합니다.
    for cur in used_currencies:
        ws = wb.create_sheet(f'환율({cur})')
        write_exchange_rate_sheet(ws, rates.get(cur))

    # ── 최종 안전장치: 불필요한 시트 삭제
    keep_sheets = {'총집계', '월별집계'}
    keep_sheets.update(cur for cur in used_currencies if cur != 'JPY')
    if qoo10_used:
        keep_sheets.update({'JPY', '큐텐(소포수령증)'})
    keep_sheets.update(_shopee_sheet_names_for_results(shopee_results, shopee_currencies))
    if _has_lazada_data(lazada_result):
        keep_sheets.add('라자다(소포수령증)')
        keep_sheets.update(f'라자다({cur})' for cur in lazada_currencies)
    keep_sheets.update(_ebay_sheet_names_for_results(ebay_results))
    keep_sheets.update(f'환율({cur})' for cur in used_currencies)
    _prune_workbook_sheets(wb, keep_sheets)

    wb.save(output_path)
    print(f'  ✅ 엑셀 저장 완료: {output_path}')
    return output_path
