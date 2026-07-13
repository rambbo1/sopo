# -*- coding: utf-8 -*-
"""영세율첨부서류제출명세서 및 수출실적명세서 생성 모듈."""

from __future__ import annotations

import re
from copy import copy
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .exchange_rate import get_rate_for_date, avg_rate_for_period, monthly_avg_rate_for_month

RATE_DIVISOR = {"JPY": 100, "VND": 100}
TRACKING_NO_PATTERN = re.compile(r"^[A-Z]{2}[A-Z0-9]{13}$", re.I)

# 수출실적명세서/통화 시트의 수출신고번호는 공란, 기타영세율건수는 1로 신고합니다.
# 영세율첨부서류제출명세서의 L/C 번호 칸에는 증빙 식별을 위해 운송장번호를 입력합니다.

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")


def safe_filename(value: str) -> str:
    value = str(value or "회사이름").strip() or "회사이름"
    for ch in '\\/:*?"<>|':
        value = value.replace(ch, "_")
    return value


def date_to_int(value):
    d = re.sub(r"\D", "", str(value or ""))[:8]
    return int(d) if len(d) == 8 else None


def date_to_month_key(value):
    d = re.sub(r"\D", "", str(value or ""))[:8]
    return f"{d[:4]}년{d[4:6]}월" if len(d) == 8 else "날짜없음"




def qoo10_reporting_date(entry=None, result=None):
    """큐텐 신고 기준일을 거래기간 종료일로 반환합니다.

    PDF/STEP 2에 거래기간 종료일이 있으면 그 날짜를 그대로 사용하고,
    종료일이 비어 있을 때만 시작일 또는 작성일을 기준으로 반기말
    (6월 30일 또는 12월 31일)을 보완합니다.
    """
    entry = entry or {}
    result = result or {}
    period_end = entry.get("period_end") or result.get("period_end") or ""
    digits = re.sub(r"\D", "", str(period_end))[:8]
    if len(digits) == 8:
        return f"{digits[:4]}-{digits[4:6]}-{digits[6:8]}"

    base = (entry.get("period_start") or result.get("period_start")
            or entry.get("write_date") or result.get("write_date") or "")
    digits = re.sub(r"\D", "", str(base))[:8]
    if len(digits) >= 6:
        year = digits[:4]
        month = int(digits[4:6])
        return f"{year}-06-30" if month <= 6 else f"{year}-12-31"
    return ""

def is_valid_tracking_no(value):
    text = re.sub(r"[^A-Za-z0-9]", "", str(value or "")).upper()
    return bool(TRACKING_NO_PATTERN.fullmatch(text))


def other_zero_rate_count_value(value=None):
    return 1


def _company_name(shopee_results, lazada_result, qoo10_result, ebay_results=None):
    for sd in shopee_results or []:
        sub = sd.get("submitter") or {}
        if sub.get("name"):
            return sub.get("name")
    if lazada_result:
        sub = lazada_result.get("submitter") or {}
        if sub.get("name"):
            return sub.get("name")
    for er in ebay_results or []:
        sub = er.get("submitter") or {}
        if sub.get("name"):
            return sub.get("name")
    if qoo10_result:
        sub = qoo10_result.get("submitter") or {}
        if sub.get("name"):
            return sub.get("name")
    return "회사이름"


def build_declaration_rows(shopee_results, lazada_result, qoo10_result, rates, ebay_results=None):
    """선택 문서 생성에 공통으로 쓰는 수출/영세율 행 목록 생성."""
    rows = []

    for sd in shopee_results or []:
        cur = sd.get("currency", "")
        div = RATE_DIVISOR.get(cur, 1)
        for tx in sd.get("transactions", []):
            rate = get_rate_for_date(rates.get(cur), tx.get("date", ""))
            amount = float(tx.get("amount", 0) or 0)
            krw = round(amount * rate / div)
            tracking = tx.get("tracking_no", "")
            ship_date = date_to_int(tx.get("date", ""))
            rows.append({
                "platform": "쇼피",
                "issuer": tx.get("carrier") or sd.get("carrier") or "주)두라로지스틱스",
                "tracking_no": tracking,
                "export_no": "",
                "other_count": 1,
                "ship_date": ship_date,
                "issue_date": ship_date,
                "currency": cur,
                "rate": rate,
                "foreign": amount,
                "krw": krw,
            })

    if lazada_result:
        ps = lazada_result.get("period_start", "")
        pe = lazada_result.get("period_end", "")
        wd = lazada_result.get("write_date", "") or pe
        ship_date = date_to_int(wd)
        for it in lazada_result.get("items", []):
            cur = it.get("currency", "")
            div = RATE_DIVISOR.get(cur, 1)
            rate = avg_rate_for_period(rates.get(cur), ps, pe)
            amount = float(it.get("amount", 0) or 0)
            krw = round(amount * rate / div)
            tracking = it.get("tracking_no", "")
            rows.append({
                "platform": "라자다",
                "issuer": it.get("carrier") or lazada_result.get("carrier") or "용성종합물류",
                "tracking_no": tracking,
                "export_no": "",
                "other_count": 1,
                "ship_date": ship_date,
                "issue_date": ship_date,
                "currency": cur,
                "rate": rate,
                "foreign": amount,
                "krw": krw,
            })


    for er in ebay_results or []:
        for it in er.get("items", []):
            cur = it.get("currency", "")
            if not cur:
                continue
            div = RATE_DIVISOR.get(cur, 1)
            rate = monthly_avg_rate_for_month(rates.get(cur), it.get("month") or it.get("date") or "")
            amount = float(it.get("amount", 0) or 0)
            krw = round(amount * rate / div)
            ship_date = date_to_int(it.get("date") or it.get("period_end") or "")
            rows.append({
                "platform": "이베이",
                "issuer": it.get("carrier") or er.get("carrier") or "린코스(주)",
                "tracking_no": it.get("tracking_no", ""),
                "export_no": "",
                "other_count": 1,
                "ship_date": ship_date,
                "issue_date": ship_date,
                "currency": cur,
                "rate": rate,
                "foreign": amount,
                "krw": krw,
            })

    if qoo10_result:
        q_entries = qoo10_result.get("entries") or [{
            "tracking_no": qoo10_result.get("tracking_no", ""),
            "qty": qoo10_result.get("qty", 0),
            "amount": qoo10_result.get("amount", 0),
            "write_date": qoo10_result.get("write_date", ""),
            "period_start": qoo10_result.get("period_start", ""),
            "period_end": qoo10_result.get("period_end", ""),
        }]
        for e in q_entries:
            ps = e.get("period_start", "") or qoo10_result.get("period_start", "")
            pe = e.get("period_end", "") or qoo10_result.get("period_end", "")
            report_date = qoo10_reporting_date(e, qoo10_result)
            rate = avg_rate_for_period(rates.get("JPY"), ps, pe)
            amount = float(e.get("amount", 0) or 0)
            krw = round(amount * rate / 100)
            tracking = e.get("tracking_no", "") or qoo10_result.get("tracking_no", "")
            ship_date = date_to_int(report_date)
            rows.append({
                "platform": "큐텐",
                "issuer": qoo10_result.get("carrier", "국제로지스틱"),
                "tracking_no": tracking,
                "export_no": "",
                "other_count": 1,
                "ship_date": ship_date,
                "issue_date": ship_date,
                "currency": "JPY",
                "rate": rate,
                "foreign": amount,
                "krw": krw,
            })

    return sorted(rows, key=lambda r: (r.get("ship_date") or 0, r.get("currency") or "", r.get("tracking_no") or ""))


def _find_template(base_dir: Path, filename: str) -> Optional[Path]:
    candidates = [
        base_dir / "forms" / filename,
        Path(__file__).resolve().parents[1] / "forms" / filename,
        base_dir / filename,
    ]
    for p in candidates:
        if p.exists():
            return p
    return None


def _copy_style(ws, src_row, dst_row, max_col):
    for col in range(1, max_col + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)


def _clear_rows(ws, start_row, max_col):
    for r in range(start_row, ws.max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(r, c).value = None


def _style_header(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row, c)
        cell.font = Font(name="맑은 고딕", size=10, bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def create_export_performance(rows, output_dir: Path, company: str, base_dir: Optional[Path] = None):
    output_dir = Path(output_dir)
    base_dir = Path(base_dir or output_dir)
    template = _find_template(base_dir, "수출실적명세서 양식.xlsx")
    if template:
        wb = load_workbook(template)
        ws = wb.active
        start_row = 2
        max_col = 7
        _clear_rows(ws, start_row, max_col)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        headers = ["수출신고번호", "기타영세율건수", "선(기)적일자", "통화코드", "환율", "외화금액", "원화금액"]
        for c, h in enumerate(headers, 1):
            ws.cell(1, c, h)
        _style_header(ws, 1, 7)
        start_row = 2
        max_col = 7

    for idx, r in enumerate(rows, start_row):
        if idx > start_row:
            _copy_style(ws, start_row, idx, max_col)
        vals = [r["export_no"], r["other_count"], r["ship_date"], r["currency"], r["rate"], r["foreign"], r["krw"]]
        for c, v in enumerate(vals, 1):
            ws.cell(idx, c, v)
        ws.cell(idx, 5).number_format = "#,##0.00"
        ws.cell(idx, 6).number_format = "#,##0.00"
        ws.cell(idx, 7).number_format = "#,##0"

    output = output_dir / f"수출실적명세서_{safe_filename(company)}.xlsx"
    wb.save(output)
    return output


def _write_zero_sheet(ws, rows):
    start_row = 3
    max_col = 15
    _clear_rows(ws, start_row, max_col)
    for idx, r in enumerate(rows, start_row):
        if idx > start_row:
            _copy_style(ws, start_row, idx, max_col)
        vals = [
            1, "소포수령증", r["issuer"], r["issue_date"], r["ship_date"], r.get("tracking_no", ""), "",
            r["currency"], r["rate"], r["foreign"], r["krw"], r["foreign"], r["krw"], 0, 0,
        ]
        for c, v in enumerate(vals, 1):
            ws.cell(idx, c, v)
        for c in [9, 10, 12, 14]:
            ws.cell(idx, c).number_format = "#,##0.00"
        for c in [11, 13, 15]:
            ws.cell(idx, c).number_format = "#,##0"


def _new_zero_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "영세율첨부서류입력"
    headers1 = ["구분", "서류명", "발급자", "발급일자", "선적일자", "L/C 번호", "비고", "통화코드", "환율", "당기 제출 금액", "", "당기 신고 해당분", "", "당기 신고 미도래 금액", ""]
    headers2 = ["", "", "", "", "", "", "", "", "", "외화", "원화", "외화", "원화", "외화", "원화"]
    for c, h in enumerate(headers1, 1):
        ws.cell(1, c, h)
    for c, h in enumerate(headers2, 1):
        ws.cell(2, c, h)
    _style_header(ws, 1, 15)
    _style_header(ws, 2, 15)
    widths = [8, 14, 18, 12, 12, 20, 12, 10, 10, 14, 14, 14, 14, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return wb


def create_zero_rate_attachments(
    rows,
    output_dir: Path,
    company: str,
    base_dir: Optional[Path] = None,
    mode: str = "both",
):
    """영세율첨부서류제출명세서 생성.

    mode:
      - "all" 또는 "전체": 전체 파일만 생성
      - "monthly" 또는 "월별": 월별 파일만 생성
      - "both" 또는 "전체+월별": 전체와 월별을 모두 생성
    """
    output_dir = Path(output_dir)
    base_dir = Path(base_dir or output_dir)
    template = _find_template(base_dir, "영세율첨부서류명세서 양식.xlsx")
    created = []

    normalized_mode = str(mode or "both").strip().lower()
    make_all = normalized_mode in {"all", "both", "전체", "전체+월별", "all+monthly"}
    make_monthly = normalized_mode in {"monthly", "both", "월별", "전체+월별", "all+monthly"}
    if not make_all and not make_monthly:
        make_all = True

    def make_file(sub_rows, suffix):
        if template:
            wb = load_workbook(template)
        else:
            wb = _new_zero_template()
        ws = wb.active
        _write_zero_sheet(ws, sub_rows)
        out = output_dir / f"영세율첨부서류제출명세서_{safe_filename(company)}_{suffix}.xlsx"
        wb.save(out)
        return out

    if make_all:
        created.append(make_file(rows, "전체"))

    if make_monthly:
        by_month = {}
        for r in rows:
            by_month.setdefault(date_to_month_key(r.get("ship_date")), []).append(r)
        for month_key in sorted(by_month):
            created.append(make_file(by_month[month_key], month_key))

    return created


def company_name_from_results(shopee_results, lazada_result=None, qoo10_result=None, ebay_results=None):
    return _company_name(shopee_results, lazada_result, qoo10_result, ebay_results=ebay_results)
